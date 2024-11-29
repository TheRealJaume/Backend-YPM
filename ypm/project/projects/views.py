# DJANGO
import os

from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from celery.result import AsyncResult
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.utils.timezone import now
from rest_framework import viewsets, status
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import action

# PROJECT
from project.projects.models import Project
from project.projects.responses import ProjectResponses
from project.projects.serializers import ProjectSerializer, ProjectListSerializer, RetrieveProjectSerializer, \
    CreateProjectSerializer, InfoProjectSerializer
from project.sprints.models import ProjectSprint
from project.task.models import ProjectTask, ProjectTaskWorker
from project.task.utils import serialize_project_tasks



class UserFilterQueryset:
    def filter_queryset(self, request, queryset, view):
        return queryset.filter(company__owner=request.user)


class ProjectViewset(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    lookup_field = 'id'
    serializer_class = ProjectSerializer
    serializer_action_classes = {
        "list": ProjectListSerializer,
        # "update": ProjectUpdateSerializer,
        "retrieve": RetrieveProjectSerializer,
        "create": CreateProjectSerializer,
        "info": InfoProjectSerializer,
    }
    filter_backends = [
        # UserRoleUserQueryset,
        UserFilterQueryset,
        SearchFilter, OrderingFilter]

    permission_classes = [
        # UserPermission,
        IsAuthenticated]

    def get_serializer_class(self):
        try:
            return self.serializer_action_classes[self.action]
        except (KeyError, AttributeError):
            super().get_serializer_class()

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        # Get serializer class
        serializer_class = self.get_serializer_class()
        serializer = serializer_class(data=request.data)
        # Check if the information sent is valid
        is_valid = serializer.validate(data=request.data)
        if is_valid:
            project = serializer.create(validated_data=request.data)
            return Response(ProjectResponses.CreateProject200({"id": project.id}), 200)
        else:
            return Response(ProjectResponses.CreateProject400(error=serializer.errors), 400)

    @action(detail=False, methods=['get'])
    def info(self, request, *args, **kwargs):
        # Get serializer class
        serializer_class = self.get_serializer_class()
        # Create event statistics
        if 'project' in request.GET:
            try:
                project = Project.objects.get(id=request.GET['project'])
            except Exception:
                return Response(ProjectResponses.DetailProject204(), 204)
            project_serializer = serializer_class(project)
            return Response(ProjectResponses.DetailProject200(project_serializer.data), 200)

    @action(detail=False, methods=['get'])
    def tasks(self, request, *args, **kwargs):
        # Create event statistics
        if 'project' in request.GET:
            try:
                project = Project.objects.get(id=request.GET['project'])
            except Exception:
                return Response(ProjectResponses.ProjectTasks204(), 204)
            project_tasks = ProjectTask.objects.filter(project=project)
            if project_tasks.count() > 0:
                serialized_tasks = serialize_project_tasks(project)
                return Response(ProjectResponses.ProjectTasks200(serialized_tasks), 200)
            else:
                return Response(ProjectResponses.ProjectTasks204(), 200)


    @action(detail=False, methods=['post'])
    def export_excel(self, request):
        try:
            # Obtain the project and its sprints
            project = Project.objects.get(id=request.data['project'])
            sprints = ProjectSprint.objects.filter(project=project).order_by('start_date')

            # Create the Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = f"Project {project.name}"

            # Write headers
            headers = ['Sprint', 'Task', 'Description', 'Estimation (h)', 'Department', 'Assigned To']
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            # Write tasks organized by sprints
            row = 2
            for sprint in sprints:
                tasks = ProjectTask.objects.filter(sprint=sprint)
                for task in tasks:
                    worker = ProjectTaskWorker.objects.filter(task=task)
                    ws.cell(row=row, column=1, value=sprint.name)
                    ws.cell(row=row, column=2, value=task.name)
                    ws.cell(row=row, column=3, value=task.description)
                    ws.cell(row=row, column=4, value=task.time)
                    ws.cell(row=row, column=5, value=task.department.department.name)
                    ws.cell(row=row, column=6,
                            value=worker.first().worker.first_name if worker.count() > 0 else 'Unassigned')
                    row += 1

            # Save the Excel file to memory
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            # Prepare the response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            filename = f"project_{project.id}_tasks_{now().strftime('%Y%m%d%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)